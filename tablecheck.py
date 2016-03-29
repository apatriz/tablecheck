#!/usr/bin/env python

"""
This script can be used to update a database table in Arc. The script
will fill in the database records with missing values based on the lookup table (
generated from an Excel .xlsx file).

"""

import arcpy,os
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string,get_column_letter
from collections import OrderedDict

###BEGIN PARAMETER INPUT#############################

# Set the Excel file to use as a lookup table
EXCEL_FILE = r"C:\Users\patrizio\Projects\Monroe_Signs\test\data_v4\Lookup_Table.xlsx"

# Set the Excel table range to extract the table header names.
# The header names should be equivalent to the field names of the database table (but do not need to have the same name)
FIELD_NAME_RANGE = "A1:S1"

# Set the lookup table column which contains the shared id (foreign key)
ID_COLUMN = "A"

# Set the database table that will be updated
DATABASE_TABLE = r"C:\Users\patrizio\Projects\Monroe_Signs\test\data_v4\Monroe_Signs.gdb\Signs"

#Set the database foreign key field name corresponding to the lookup table id
FOREIGN_KEY = "SignType"

# Set the field map.
# Maps the field name in the database table to the header in the lookup table. This
# is used in case the field naming convention/order differs between both tables.
FIELD_MAP = {"SignType":"Code",
             "DimHeight":"DimHeight",
             "DimWidth":"DimWidth",
             "LegendColor1":"LegendColor1",
             "LegendColor2":"LegendColor2",
             "LegendColor3":"LegendColor3",
             "SheetColor1":"SheetingColor1",
             "SheetColor2":"SheetingColor2",
             "RegPkRestr1":"RegPkRestrType1",
             "RegPkRestr2":"RegPkRestrType2",
             "RegPkTimeLim1":"RegPkTimeLimit1",
             "RegPkTimeLim2":"RegPkTimeLimit2",
             "RegPkArrow1":"RegPkArrow1",
             "RegPkTimeYear1":"RegPkTimeYear1",
             "RegPkTimeYear2":"RegPkTimeYear2",
             "RegPkVehExcep1":"RegPkVehExceptions1",
             "RegPkVehExcep2":"RegPkVehExceptions2",
             }

##############END PARAMETER INPUT############################


#decorator to convert indexes to 0-based
def convertindex(func):
    def minus_one(index_string):
        return func(index_string) - 1
    return minus_one


def get_headers(filename,field_range):
    """
    Input:
    filename -> path to excel file
    field_range -> range of cell coordinates for getting headers, e.g. 'A1:S1'

    Output: list
    """
    wb = load_workbook(filename=filename,read_only=True)
    ws = wb.active
    headers = []

    for row in ws.iter_rows(range_string=field_range):
        for cell in row:
            headers.append(cell.value)
    return headers


def load_dict(filename,field_range,id_column):
    '''Load a dictionary with the lookup table from excel file'''
    wb = load_workbook(filename=filename,read_only=True)
    ws = wb.active
    res = {}
    get_index = convertindex(column_index_from_string)
    headers = get_headers(filename,field_range)
    for row in ws.iter_rows(row_offset=1):
        index = 0
        entry = {}
        for cell in row:
            if index <= len(headers) -1:
                try:
                    val = cell.value.strip()
                except AttributeError:
                    val = cell.value
                entry[headers[index]] = val
                index += 1
        res[row[get_index(id_column)].value] = entry
    return res


def convert_null_to_none(filename,field_map):
    '''Converts all values with "Null" in the string to None'''
    fields = [field.name for field in arcpy.ListFields(filename) if field.name in field_map]
    with arcpy.da.UpdateCursor(filename,fields) as cursor:
        for row in cursor:
            new_row = []
            for cell in row:
                if "Null" in str(cell):
                    cell = None
                else:
                    cell = cell
                new_row.append(cell)
            cursor.updateRow(new_row)
    print "Converted null strings to true None values"


def get_exception_category_list(row_dict,exception_field):
    '''Get list of values for category of fields'''
    result = []
    for field in row_dict:
        if exception_field in str(field):
            result.append(row_dict[field])
    return result

        
def fill_cell(value_list):
    '''Returns True if all values in the list are None'''
    for i in value_list:
        if i != None:
            return False
    return True


def lookup_value(lookup_id, lookup_table, fieldname):
    '''Gets the cell value from a lookup table based on a lookup id and field'''

    lookup_field = fieldname       
    # check if the lookup table contains a record for that id and field
    if lookup_table[lookup_id][lookup_field]:
        # return the lookup table value
        return lookup_table[lookup_id][lookup_field]

def convert_field_name(field_name, field_map,reverse=False):
    '''Converts field name based on field map. Accomodates reverse lookup.'''
    if reverse:
        field_map = {v:k for k,v in field_map.items()}
    return field_map[field_name]
                
# TODO: Get rid of hard-coded strings SheetColor and LegendColor
def update_arc_table(filename,foreign_key,field_map,lookup_table):
    fields = [field.name for field in arcpy.ListFields(filename) if field.name in field_map]
    with arcpy.da.UpdateCursor(filename,fields) as cursor:
        for row in cursor:
            # initialize index to keep track of field position
            index = 0
            # initialize the row to hold the row values
            new_row = OrderedDict()
            sheeting_color = []
            legend_color = []
            for cell in row:
                new_row[fields[index]] = cell
                index += 1

            # check if foreign_key field exists and it has a value
            if foreign_key in new_row and new_row[foreign_key]:
                lookup_id = new_row[foreign_key]
            else:
                # if it doesn't, continue to next row
                continue
            
            # get exception category values
            sheet_color = get_exception_category_list(new_row,"SheetColor")
            legend_color = get_exception_category_list(new_row, "LegendColor")

            # check if exceptions should be filled
            fill_sheetcolor = fill_cell(sheet_color)
            fill_legendcolor = fill_cell(legend_color)
       
            # loop through each record in the row, updating where appropriate
            for field in new_row:
                if "SheetColor" in str(field) and not fill_sheetcolor:
                        continue
                elif "LegendColor" in str(field) and not fill_legendcolor:
                        continue
                else:
                    # when finding a blank cell, and the lookup id (foreign key) is in the lookup table,
                    if not new_row[field] and lookup_id in lookup_table:
                        # convert field name
                        lookup_field_name = convert_field_name(field,field_map)
                        # set the new cell value to the matching lookup table value
                        cell = lookup_value(lookup_id, lookup_table, lookup_field_name)
                        new_row[field] = cell
                 
            # create new row from ordered dict values
            new_row = tuple(new_row.values())
            cursor.updateRow(new_row)
            
    print "Finsihed updating table {0}".format(os.path.abspath(filename))
    return os.path.abspath(filename)


### TESTS ###
def test_load_dict(test_entries):
    
    lookup_table = load_dict(EXCEL_FILE,FIELD_NAME_RANGE,ID_COLUMN)
    
    for i in test_entries:
        assert (lookup_table[i] == test_entries[i]),"{0} DOES NOT EQUAL {1}".format(lookup_table[i],test_entries[i]) 
    print "Test complete."


if __name__ == "__main__":
    
    convert_null_to_none(DATABASE_TABLE,FIELD_MAP)
    lookup_table = load_dict(EXCEL_FILE,FIELD_NAME_RANGE,ID_COLUMN)
    update_arc_table(DATABASE_TABLE,FOREIGN_KEY,FIELD_MAP,lookup_table)

    

##    test_load_dict({"D4-3":{u'Code': u'D4-3', u'SheetingColor2': None,
##                                        u'RegPkVehExceptions1': None, u'DimHeight': 18L,
##                                        u'SheetingColor1': u'White', u'LegendColor1': u'Green',
##                                        u'LegendColor2': None, u'LegendColor3': None, u'RegPkRestrType2': None,
##                                        u'RegPkRestrType1': None, u'RegPkTimeLimit2': None, u'Collect': u'LiDAR',
##                                        u'RegPkTimeLimit1': None, u'RegPkVehExceptions2': None, u'RegPkArrow1': None,
##                                        u'Descrip': u'Bike Parking: D4-3', u'RegPkTimeYear1': None, u'RegPkTimeYear2': None, u'DimWidth': 12L},
##                          "W1-2aL":{u'Code': u'W1-2aL', u'SheetingColor2': None, u'RegPkVehExceptions1': None,
##                                     u'DimHeight': 30L, u'SheetingColor1': u'Yellow', u'LegendColor1': u'Black',
##                                     u'LegendColor2': None, u'LegendColor3': None, u'RegPkRestrType2': None,
##                                     u'RegPkRestrType1': None, u'RegPkTimeLimit2': None, u'Collect': u'LiDAR',
##                                     u'RegPkTimeLimit1': None, u'RegPkVehExceptions2': None,
##                                     u'RegPkArrow1': None, u'Descrip': u'Curve Left: W1-2aL',
##                                     u'RegPkTimeYear1': None, u'RegPkTimeYear2': None, u'DimWidth': 30L}})



